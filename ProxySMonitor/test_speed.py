import datetime
import os
import subprocess
import openpyxl

def test_speed(subscription_url, proxy_name, output_dir):
    current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(output_dir, f"{proxy_name}_{current_time}.txt")
    
    # 设置工作目录为 ssrspeedn 所在目录
    os.chdir('ssrspeedn')
    
    command = f"python3 main.py -u '{subscription_url}' -g {proxy_name} -M pingonly --skip-requirements-check"
    
    with open(output_file, 'w') as f:
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        
        while True:
            output = process.stdout.readline() if process.stdout else None
            if output == '' and process.poll() is not None:
                break
            if output:
                print(output.strip())
                f.write(output)
                f.flush()
                
        stderr = process.stderr.read() if process.stderr else None
        if stderr:
            print(stderr.strip())
            f.write(stderr)
            f.flush()
    
    # 切换回原来的工作目录
    os.chdir('..')
    
    print(f"Test result saved to {output_file}")

if __name__ == "__main__":
    folder_path = './'
    workbook = openpyxl.load_workbook('StabilityM.xlsx')  # 返回一个workbook数据类型的值
    sheet = workbook['Sheet1']     # 获取活动表

    rows = sheet.max_row        # 获取行数
    for i in range(1, rows + 1): 
        proxy_name = str(sheet.cell(row=i, column=1).value)       # 获取第i行第1列的数据
        subscription_url = str(sheet.cell(row=i, column=2).value) 
        output_dir = "./results"  # 使用相对路径保存结果
        os.makedirs(output_dir, exist_ok=True)  # 如果文件夹不存在则创建
        test_speed(subscription_url, proxy_name, output_dir)
