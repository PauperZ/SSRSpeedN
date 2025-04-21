import json
from PIL import Image, ImageDraw, ImageFont
import os
from datetime import datetime

import openpyxl

def calculate_online_rate(data):
    """计算节点在线率"""
    online_rates = []
    for node in data:
        total_pings = len(node["pingRecord"]["gPings"])
        successful_pings = node["pingRecord"]["ping_suc"]
        online_rate = (successful_pings / total_pings) * 100 if total_pings > 0 else 0
        online_rates.append(online_rate)
    return online_rates

def export_images(folder_path):
    os.chdir(folder_path)
    workbook = openpyxl.load_workbook('StabilityM.xlsx')  # 返回一个workbook数据类型的值
    sheet = workbook['Sheet1']     # 获取活动表

    rows = sheet.max_row        # 获取行数
    ping_res_path = folder_path+'/results'
    for i in range(rows): 
        proxy_name = str(sheet.cell(row=i+1,column=1).value)       # 获取第i行第1列的数据

        files = os.listdir(ping_res_path)
        for file_name in files: 
            # 构造完整文件路径
            file_path = os.path.join(ping_res_path, file_name)  

            # 只处理以 .json 结尾且以代理名字开头的文件
            if file_name.endswith('.json') and file_name.startswith(proxy_name):
                with open(file_path, 'r',encoding='utf-8') as json_file:
                    single_proxy_res_data = json.load(json_file)
                    gen_image(single_proxy_res_data,proxy_name)


def gen_image(single_proxy_res_data,proxy_name):
    data = single_proxy_res_data
    
    # 计算节点在线率
    online_rates = calculate_online_rate(data)

    # 设置字体（使用支持中文的字体）
    font_path = "PingFang Medium.ttf"  # 替换为你的字体文件路径
    font = ImageFont.truetype(font_path, 20)
    small_font = ImageFont.truetype(font_path, 14)  # 小字体用于左上角格子

    # 创建空白图像
    padding = 10  # 单元格内文本的填充
    title_height = 40  # 新增行高度
    header_height = 40  # 表头行高度
    footer_height = 40  # 表尾行高度

    # 计算图像的大小
    rows = len(data)
    columns = max(len(node["pingRecord"]["gPings"]) for node in data) + 1  # 增加一列用于显示在线率

    # 计算表格的列宽度和行高度
    max_widths = [0] * (columns + 1)
    max_heights = [header_height] * (rows + 1) + [footer_height]

    # 计算最大列宽度和行高度
    for i, node in enumerate(data):
        node_name = node["node_name"]
        text_width, text_height = font.getbbox(node_name)[2:4]
        max_widths[0] = max(max_widths[0], text_width + padding * 2)
        max_heights[i + 1] = max(max_heights[i + 1], text_height + padding * 2)
        for j, ping in enumerate(node["pingRecord"]["gPings"]):
            gPing = int(ping["gPing"] * 1000)  # 转换为毫秒并取整
            gPing_text = str(gPing)
            text_width, text_height = font.getbbox(gPing_text)[2:4]
            max_widths[j + 1] = max(max_widths[j + 1], text_width + padding * 2)
            max_heights[i + 1] = max(max_heights[i + 1], text_height + padding * 2)

        # 计算在线率列的宽度
        online_rate_text = f"{online_rates[i]:.2f}%"
        text_width, text_height = font.getbbox(online_rate_text)[2:4]
        max_widths[-1] = max(max_widths[-1], text_width + padding * 2)

    # 计算表头列宽度
    for j in range(columns - 1):  # 不包括在线率列
        test_time = data[0]["pingRecord"]["gPings"][j]["test_time"]
        formatted_time = datetime.strptime(test_time, "%Y-%m-%d-%H-%M-%S").strftime("%Y-%m-%d %H:%M:%S")
        text_width, text_height = font.getbbox(formatted_time)[2:4]
        max_widths[j + 1] = max(max_widths[j + 1], text_width + padding * 2)

    # 计算图像尺寸
    image_width = sum(max_widths)
    image_height = sum(max_heights) + title_height

    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)

    # 绘制标题行
    light_gray = (200, 200, 200)  # 浅色颜色
    draw.rectangle((0, 0, image_width, title_height), outline=light_gray)
    title_text = "便宜机场测速-机场稳定性分析"
    title_text_width, title_text_height = font.getbbox(title_text)[2:4]
    title_x = (image_width - title_text_width) / 2
    title_y = (title_height - title_text_height) / 2
    draw.text((title_x, title_y), title_text, fill="black", font=font)

    # 绘制表头
    for j in range(columns - 1):  # 不包括在线率列
        test_time = data[0]["pingRecord"]["gPings"][j]["test_time"]
        formatted_time = datetime.strptime(test_time, "%Y-%m-%d-%H-%M-%S").strftime("%Y-%m-%d %H:%M:%S")
        text_width, text_height = font.getbbox(formatted_time)[2:4]
        x = sum(max_widths[:j + 1]) + (max_widths[j + 1] - text_width) / 2
        y = title_height + (header_height - text_height) / 2
        draw.rectangle((sum(max_widths[:j + 1]), title_height, sum(max_widths[:j + 2]), title_height + header_height), outline=light_gray)
        draw.text((x, y), formatted_time, fill="black", font=font)

    # 绘制在线率表头
    online_rate_header = "节点在线率"
    text_width, text_height = font.getbbox(online_rate_header)[2:4]
    x = sum(max_widths[:-1])
    y = title_height + (header_height - text_height) / 2
    draw.rectangle((x, title_height, image_width, title_height + header_height), outline=light_gray)
    draw.text((x + (max_widths[-1] - text_width) / 2, y), online_rate_header, fill="black", font=font)

    # 绘制左上角空白格子，并添加斜线和文字
    draw.rectangle((0, title_height, max_widths[0], title_height + header_height), outline=light_gray)
    draw.line((0, title_height, max_widths[0], title_height + header_height), fill=light_gray)

    # 在左下角绘制“NodeName”
    node_text = "NodeName"
    node_text_width, node_text_height = small_font.getbbox(node_text)[2:4]
    node_x = padding
    node_y = title_height + header_height - node_text_height - padding
    draw.text((node_x, node_y), node_text, fill="black", font=small_font)

    # 在右上角绘制“Time”
    time_text = "Time"
    time_text_width, time_text_height = small_font.getbbox(time_text)[2:4]
    time_x = max_widths[0] - time_text_width - padding
    time_y = title_height + padding
    draw.text((time_x, time_y), time_text, fill="black", font=small_font)

    # 绘制左侧节点名称和表格内容
    for i, node in enumerate(data):
        node_name = node["node_name"]
        text_width, text_height = font.getbbox(node_name)[2:4]
        x = (max_widths[0] - text_width) / 2
        y = title_height + sum(max_heights[:i + 1]) + (max_heights[i + 1] - text_height) / 2
        draw.rectangle((0, title_height + sum(max_heights[:i + 1]), max_widths[0], title_height + sum(max_heights[:i + 2])), outline=light_gray)
        draw.text((x, y), node_name, fill="black", font=font)
        for j in range(columns - 1):  # 不包括在线率列
            if j < len(node["pingRecord"]["gPings"]):
                ping = node["pingRecord"]["gPings"][j]
                gPing = int(ping["gPing"] * 1000)  # 转换为毫秒并取整
                gPing_text = str(gPing) + "ms"
                text_width, text_height = font.getbbox(gPing_text)[2:4]
                cell_x = sum(max_widths[:j + 1])
                cell_y = title_height + sum(max_heights[:i + 1])
                
                if gPing == 0:
                    fill_color = (255, 182, 193)  # Light red for gPing == 0
                else:
                    if gPing <= 100:
                        fill_color = (127, 255, 170)  # Light blue
                    elif gPing <= 500:
                        fill_color = (187, 255, 255)  # Sky blue
                    elif gPing <= 1000:
                        fill_color = (152, 245, 255)  # Steel blue
                    elif gPing <= 1500:
                        fill_color = (142, 229, 238)  # Deep sky blue
                    else:
                        fill_color = (122, 197, 205)  # Midnight blue
            else:
                gPing_text = "N/A"
                text_width, text_height = font.getbbox(gPing_text)[2:4]
                cell_x = sum(max_widths[:j + 1])
                cell_y = title_height + sum(max_heights[:i + 1])
                fill_color = "white"

            draw.rectangle((cell_x, cell_y, cell_x + max_widths[j + 1], cell_y + max_heights[i + 1]), fill=fill_color, outline=light_gray)
            draw.text((cell_x + (max_widths[j + 1] - text_width) / 2, cell_y + (max_heights[i + 1] - text_height) / 2), gPing_text, fill="black", font=font)

        # 绘制在线率
        online_rate_text = f"{online_rates[i]:.2f}%"
        text_width, text_height = font.getbbox(online_rate_text)[2:4]
        cell_x = sum(max_widths[:-1])
        cell_y = title_height + sum(max_heights[:i + 1])
        draw.rectangle((cell_x, cell_y, image_width, cell_y + max_heights[i + 1]), outline=light_gray)
        draw.text((cell_x + (max_widths[-1] - text_width) / 2, cell_y + (max_heights[i + 1] - text_height) / 2), online_rate_text, fill="black", font=font)

    # 绘制表尾
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    footer_text = "测速频道：TG@Cheap_Proxy 作者：TG@winderosion        Generated at:" + current_time
    text_width, text_height = font.getbbox(footer_text)[2:4]
    x = (image_width - text_width) / 2
    y = image_height - footer_height + (footer_height - text_height) / 2
    draw.text((x, y), footer_text, fill="black", font=font)

    # 保存图像
    os.makedirs('./results', exist_ok=True)
    image.save(f"./results/{proxy_name}.png")
