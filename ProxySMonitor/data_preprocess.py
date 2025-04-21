import os
import json

import openpyxl

from gen_results.a_proxy_mul_ping_gen_results import export_images

def read_json_files(folder_path,proxy_name):
    #一个机场在一段时间内数次测速的结果，被处理成正确的的格式后返回
    res_of_one_proxy = []
    # 获取文件夹中的所有文件名
    files = os.listdir(folder_path)

    num = 0
    node_names = set()
    #第一遍循环确定每个节点的名字
    for file_name in files:
        # 构造完整文件路径
        file_path = os.path.join(folder_path, file_name)
        
        # 只处理以 .json 结尾且以代理名字开头的文件
        if file_name.endswith('.json') and file_name.startswith(proxy_name):
            with open(file_path, 'r') as json_file:
                try:
                    num += 1
                    
                    # data是一次测速的结果
                    data = json.load(json_file)

                    for single_node in data:
                        #node["node_name"] = single_node["remarks"]
                        node_names.add(single_node["remarks"])
                except json.JSONDecodeError:
                    print(f"Error decoding JSON from file: {file_path}") 

    for node_name_ele in iter(node_names):
        node = {
                    "node_name":None,
                    "pingRecord":{
                        "ping_times": 0,
                        "ping_suc":0,
                        "gPings": []
                    }
                }
        node["node_name"] = node_name_ele
        res_of_one_proxy.append(node)
          
    
    
    num = 0
    # 遍历每个文件
    for file_name in files:
        # 构造完整文件路径
        file_path = os.path.join(folder_path, file_name)
        
        # 只处理以 .json 结尾且以代理名字开头的文件
        if file_name.endswith('.json') and file_name.startswith(proxy_name):
            with open(file_path, 'r') as json_file:
                try:
                    # data是一次测速的结果
                    data = json.load(json_file)
                    test_time = file_name[len(proxy_name):-5]
                    num = 0
                    #遍历当前打开的json文件中的所有节点
                    for node_of_data in data:
                        cur_node_name = node_of_data["remarks"]      
                        record = {
                            "test_time":None,
                            "gPing": None
                        }

                        #去匹配当前打开的json文件里的节点与结果里的节点
                        for cur_res_node in res_of_one_proxy:
                            if(cur_res_node["node_name"] == cur_node_name):
                                num += 1
                                record["gPing"] = node_of_data["gPing"]
                                record["test_time"] = test_time
                                cur_res_node["pingRecord"]["gPings"].append(record)
                                cur_res_node["pingRecord"]["ping_times"] += 1
                                if(node_of_data["gPing"] != 0):
                                    cur_res_node["pingRecord"]["ping_suc"] += 1
                    
                except json.JSONDecodeError:
                    print(f"Error decoding JSON from file: {file_path}")
    
    # 按数据个数排序，数据个数越少的，排序越往后
    res_of_one_proxy.sort(key=lambda x: x["pingRecord"]["ping_times"], reverse=True)
    return res_of_one_proxy

def build_result_json(folder_path):
    print("执行了build_result_json")
    os.chdir(folder_path)
    workbook = openpyxl.load_workbook('StabilityM.xlsx')  # 返回一个workbook数据类型的值
    sheet = workbook['Sheet1']     # 获取活动表

    rows = sheet.max_row        # 获取行数
    for i in range(rows): 
        proxy_name = str(sheet.cell(row=i+1,column=1).value)       # 获取第i行第1列的数据
        single_json = read_json_files(folder_path + '/ping_results',proxy_name)
        print(single_json)
        
        # 确保 results 文件夹存在
        results_folder_path = "./results"
        if not os.path.exists(results_folder_path):
            os.makedirs(results_folder_path)
        single_json_file_name = os.path.join(results_folder_path, f"{proxy_name}_result.json")  # 定义要保存的文件名
       
        with open(single_json_file_name, 'w', encoding='utf-8') as f:
            json.dump(single_json, f, ensure_ascii=False, indent=4)  # 保存 JSON 文件
        

if (__name__ == "__main__"):
    print("执行此语句")
    folder_path = './'
    build_result_json(folder_path)
    export_images(folder_path)

